# processor.py
# Lógica de negocio: operaciones sobre Excel
import re
from openpyxl import load_workbook


# Misión 1 - Procesamiento de Excel

def clean_id(value):
    # Elimina caracteres no numéricos de un documento
    if value is None:
        return ""   
    # si en value no hay nada, retorna una cadena vacía
    return re.sub(r'\D','', str(value)) # Reemplaza todo lo que no sea dígito con una cadena vacía
    # str es transformar a cadena de texto

# ==========================
#  FUNCION merge_name
#  Une nombre y apellido en un solo campo
# ==========================

def merge_name(name, last_name):
    if name is None:
        name = ""
    if last_name is None:
        last_name = ""
    return f"{name} {last_name}".strip()
    # strip elimina espacios en blanco al inicio y al final

def process_excel(path):
    # Acceso a la hoja llamada "Datos"
    wb=load_workbook(path)
    ws = wb["Datos"]
    # Recorrer todas las filas desde la fila 
    for row in range(2, ws.max_row + 1):
        # row es la fila actual en el bucle
        # le estoy diciendo que inicie desde la fila 2 hasta la última fila (ws.max_row)
        # ws.max_row es el número total de filas en la hoja de cálculo
        ws[f"D{row}"] = clean_id(ws[f"A{row}"].value)
        # En la columna D va a colocar lo que hay en la columna A pero limpio.
        ws[f"E{row}"] = merge_name(
        ws[f"B{row}"].value,
        ws[f"C{row}"].value
        )   
    wb.save(path) # Guarda los cambios en el archivo de Excel



# CONTROLADOR DE ERRORES 
def process_excel_safe(path): # Esta función es para manejar errores y retornar mensajes
    try: # Si esos se revientan, captura el error
    # try es para intentar ejecutar un bloque de código
        process_excel(path)
        return True, "Archivo procesado correctamente" # Pyhton me permite retornar más de un parámetro, en este caso un bool y un string
    except PermissionError: # Captura el error de permiso denegado cuando tengo el archivo abierto, el típico aviso de "El archivo está en uso"
        return(
            False,
            "El archivo de Excel está abierto\n." 
            "Por favor, ciérrelo e intente de nuevo"
        )
    except KeyError:
        return False, "Hoja 'Datos' no encontrada" # Las comillas '' van a aparecer en el mensaje.
    except Exception as e: # Captura cualquier otro error que no haya sido capturado antes
        # e es un alias para el objeto de excepción que contiene información sobre el error, para abreviar el Exception
        return False, f"Ocurrió un error inesperado: {str(e)}"







# Misión 2 - Procesamiento de instrucciones en lenguaje natural

def ejecutar_accion(instruccion, path):
    # Abre el archivo de ejemplo
    wb = load_workbook(path)
    ws = wb.active

    if instruccion["action"] == "clean_id":
        col = instruccion["column"]
        for fila in range(2, ws.max_row + 1):
            ws[f"{col}{fila}"] = ''.join(filter(str.isdigit, str(ws[f"{col}{fila}"].value)))

    elif instruccion["action"] == "merge_name":
        for fila in range(2, ws.max_row + 1):
            nombre = ws["A" + str(fila)].value or ""
            apellido = ws["B" + str(fila)].value or ""
            ws["C" + str(fila)] = f"{nombre} {apellido}".strip()

    wb.save(path)


